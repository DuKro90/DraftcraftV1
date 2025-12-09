"""Unit tests for bulk upload service.

Tests cover:
- Excel and CSV parsing
- German number format parsing
- Data validation
- Error handling
- Dry-run mode
- Update existing entries
"""

import io
from decimal import Decimal
from datetime import date
from django.test import TestCase
from django.contrib.auth.models import User

import openpyxl

from documents.services.bulk_upload_service import (
    BulkUploadService,
    GermanNumberParser,
    BulkUploadResult,
    ValidationError,
)
from documents.services.template_generator import TemplateGenerator
from documents.betriebskennzahl_models import (
    BetriebskennzahlTemplate,
    HolzartKennzahl,
    OberflächenbearbeitungKennzahl,
    KomplexitaetKennzahl,
    MateriallistePosition,
    SaisonaleMarge,
)


class GermanNumberParserTestCase(TestCase):
    """Test German number format parsing."""

    def setUp(self):
        self.parser = GermanNumberParser()

    def test_parse_german_decimal(self):
        """Test parsing German decimal format (1.234,56)."""
        self.assertEqual(self.parser.parse_decimal('1.234,56'), Decimal('1234.56'))
        self.assertEqual(self.parser.parse_decimal('1234,56'), Decimal('1234.56'))
        self.assertEqual(self.parser.parse_decimal('0,5'), Decimal('0.5'))

    def test_parse_us_decimal(self):
        """Test parsing US decimal format (1234.56)."""
        self.assertEqual(self.parser.parse_decimal('1234.56'), Decimal('1234.56'))
        self.assertEqual(self.parser.parse_decimal('0.5'), Decimal('0.5'))

    def test_parse_integer(self):
        """Test parsing integer."""
        self.assertEqual(self.parser.parse_decimal('1234'), Decimal('1234'))

    def test_parse_invalid_decimal(self):
        """Test parsing invalid decimal."""
        with self.assertRaises(ValueError):
            self.parser.parse_decimal('abc')
        with self.assertRaises(ValueError):
            self.parser.parse_decimal('')

    def test_parse_german_date(self):
        """Test parsing German date format (DD.MM.YYYY)."""
        self.assertEqual(
            self.parser.parse_date('31.12.2025'),
            date(2025, 12, 31)
        )

    def test_parse_iso_date(self):
        """Test parsing ISO date format (YYYY-MM-DD)."""
        self.assertEqual(
            self.parser.parse_date('2025-12-31'),
            date(2025, 12, 31)
        )

    def test_parse_invalid_date(self):
        """Test parsing invalid date."""
        with self.assertRaises(ValueError):
            self.parser.parse_date('invalid')


class HolzartBulkUploadTestCase(TestCase):
    """Test bulk upload for HolzartKennzahl."""

    def setUp(self):
        """Set up test fixtures."""
        self.user = User.objects.create_user(username='testuser', password='testpass')
        self.template = BetriebskennzahlTemplate.objects.create(
            name='Test Template',
            version='1.0',
            is_active=True,
            created_by=self.user
        )
        self.service = BulkUploadService(user=self.user)

    def _create_excel_file(self, data):
        """Helper to create Excel file bytes."""
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write data
        for row_data in data:
            ws.append(row_data)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def test_upload_holzart_valid_excel(self):
        """Test successful upload of valid Holzart data from Excel."""
        data = [
            ['holzart', 'kategorie', 'preis_faktor', 'verfuegbarkeit', 'is_enabled'],
            ['Eiche', 'hartholz', '1,3', 'Verfügbar', 'true'],
            ['Buche', 'hartholz', '1,2', 'Standard', 'true'],
        ]

        file_content = self._create_excel_file(data)

        result = self.service.upload_holzart_kennzahlen(
            file_content=file_content,
            template_id=self.template.id,
            file_format='xlsx',
            dry_run=False
        )

        self.assertTrue(result.success)
        self.assertEqual(result.created_count, 2)
        self.assertEqual(result.errors, [])

        # Verify database
        self.assertEqual(HolzartKennzahl.objects.filter(template=self.template).count(), 2)

        eiche = HolzartKennzahl.objects.get(holzart='Eiche')
        self.assertEqual(eiche.kategorie, 'hartholz')
        self.assertEqual(eiche.preis_faktor, Decimal('1.3'))
        self.assertTrue(eiche.is_enabled)

    def test_upload_holzart_dry_run(self):
        """Test dry-run mode doesn't save data."""
        data = [
            ['holzart', 'kategorie', 'preis_faktor', 'verfuegbarkeit', 'is_enabled'],
            ['Eiche', 'hartholz', '1,3', 'Verfügbar', 'true'],
        ]

        file_content = self._create_excel_file(data)

        result = self.service.upload_holzart_kennzahlen(
            file_content=file_content,
            template_id=self.template.id,
            file_format='xlsx',
            dry_run=True
        )

        self.assertTrue(result.success)
        self.assertEqual(result.created_count, 1)

        # Verify nothing saved
        self.assertEqual(HolzartKennzahl.objects.count(), 0)

    def test_upload_holzart_invalid_kategorie(self):
        """Test validation error for invalid kategorie."""
        data = [
            ['holzart', 'kategorie', 'preis_faktor', 'verfuegbarkeit', 'is_enabled'],
            ['Eiche', 'invalid', '1,3', 'Verfügbar', 'true'],
        ]

        file_content = self._create_excel_file(data)

        result = self.service.upload_holzart_kennzahlen(
            file_content=file_content,
            template_id=self.template.id,
            file_format='xlsx',
            dry_run=False
        )

        self.assertEqual(len(result.errors), 1)
        error = result.errors[0]
        self.assertEqual(error.field, 'kategorie')
        self.assertIn('hartholz, weichholz, or nadelholz', error.error)

    def test_upload_holzart_update_existing(self):
        """Test updating existing entries."""
        # Create initial entry
        HolzartKennzahl.objects.create(
            template=self.template,
            holzart='Eiche',
            kategorie='hartholz',
            preis_faktor=Decimal('1.2'),
            is_enabled=True
        )

        # Upload with updated price
        data = [
            ['holzart', 'kategorie', 'preis_faktor', 'verfuegbarkeit', 'is_enabled'],
            ['Eiche', 'hartholz', '1,5', 'Verfügbar', 'true'],
        ]

        file_content = self._create_excel_file(data)

        result = self.service.upload_holzart_kennzahlen(
            file_content=file_content,
            template_id=self.template.id,
            file_format='xlsx',
            dry_run=False,
            update_existing=True
        )

        self.assertTrue(result.success)
        self.assertEqual(result.updated_count, 1)
        self.assertEqual(result.created_count, 0)

        # Verify update
        eiche = HolzartKennzahl.objects.get(holzart='Eiche')
        self.assertEqual(eiche.preis_faktor, Decimal('1.5'))


class MateriallisteBulkUploadTestCase(TestCase):
    """Test bulk upload for MateriallistePosition."""

    def setUp(self):
        """Set up test fixtures."""
        self.user = User.objects.create_user(username='testuser', password='testpass')
        self.service = BulkUploadService(user=self.user)

    def _create_excel_file(self, data):
        """Helper to create Excel file bytes."""
        wb = openpyxl.Workbook()
        ws = wb.active

        for row_data in data:
            ws.append(row_data)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def test_upload_material_valid(self):
        """Test successful upload of valid material data."""
        data = [
            [
                'material_name', 'sku', 'lieferant', 'standardkosten_eur',
                'verpackungseinheit', 'verfuegbarkeit', 'rabatt_ab_100',
                'rabatt_ab_500', 'is_enabled'
            ],
            [
                'Schrauben 4x40mm', 'SCH-001', 'Würth', '12,50',
                'Box 100', 'Auf Lager', '5', '10', 'true'
            ],
        ]

        file_content = self._create_excel_file(data)

        result = self.service.upload_materialliste(
            file_content=file_content,
            user=self.user,
            file_format='xlsx',
            dry_run=False
        )

        self.assertTrue(result.success)
        self.assertEqual(result.created_count, 1)

        # Verify database
        material = MateriallistePosition.objects.get(sku='SCH-001')
        self.assertEqual(material.material_name, 'Schrauben 4x40mm')
        self.assertEqual(material.standardkosten_eur, Decimal('12.50'))
        self.assertEqual(material.rabatt_ab_100, Decimal('5'))

    def test_upload_material_duplicate_sku(self):
        """Test handling of duplicate SKU."""
        # Create initial material
        MateriallistePosition.objects.create(
            user=self.user,
            material_name='Test Material',
            sku='SCH-001',
            lieferant='Supplier',
            standardkosten_eur=Decimal('10'),
            is_enabled=True
        )

        data = [
            [
                'material_name', 'sku', 'lieferant', 'standardkosten_eur',
                'verpackungseinheit', 'verfuegbarkeit', 'rabatt_ab_100',
                'rabatt_ab_500', 'is_enabled'
            ],
            [
                'New Material', 'SCH-001', 'Würth', '12,50',
                'Box 100', 'Auf Lager', '5', '10', 'true'
            ],
        ]

        file_content = self._create_excel_file(data)

        # Without update_existing, should skip
        result = self.service.upload_materialliste(
            file_content=file_content,
            user=self.user,
            file_format='xlsx',
            dry_run=False,
            update_existing=False
        )

        self.assertEqual(result.skipped_count, 1)
        self.assertEqual(result.created_count, 0)


class TemplateGeneratorTestCase(TestCase):
    """Test template generator."""

    def setUp(self):
        self.generator = TemplateGenerator()

    def test_generate_holzart_excel_template(self):
        """Test Excel template generation for Holzart."""
        file_bytes = self.generator.generate_holzart_template(file_format='xlsx')

        # Load and verify
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active

        # Check headers
        headers = [cell.value for cell in ws[1]]
        self.assertIn('holzart', headers)
        self.assertIn('kategorie', headers)
        self.assertIn('preis_faktor', headers)

        # Check example data exists
        self.assertGreater(ws.max_row, 1)

    def test_generate_material_csv_template(self):
        """Test CSV template generation for Material."""
        file_bytes = self.generator.generate_materialliste_template(file_format='csv')

        # Decode and check content
        content = file_bytes.decode('utf-8-sig')
        self.assertIn('material_name', content)
        self.assertIn('sku', content)
        self.assertIn('lieferant', content)
