"""Template generator for bulk upload Excel/CSV templates.

Generates downloadable Excel templates with:
- Pre-filled headers
- Example data rows
- Data validation rules
- German formatting
"""

import io
from typing import List, Dict, Any
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import csv


class TemplateGenerator:
    """Generate Excel/CSV templates for bulk upload."""

    # Color scheme for templates
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT = Font(color='FFFFFF', bold=True)
    EXAMPLE_FILL = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

    # =========================================================================
    # HOLZART (WOOD TYPES) TEMPLATE
    # =========================================================================

    def generate_holzart_template(self, file_format: str = 'xlsx') -> bytes:
        """
        Generate template for HolzartKennzahl bulk upload.

        Columns:
        - holzart: Wood type name
        - kategorie: hartholz/weichholz/nadelholz
        - preis_faktor: Price multiplier (German format: 1,3)
        - verfuegbarkeit: Standard/Verfügbar/Knapp/Nicht verfügbar
        - is_enabled: true/false

        Returns:
            File bytes (Excel or CSV)
        """
        headers = ['holzart', 'kategorie', 'preis_faktor', 'verfuegbarkeit', 'is_enabled']

        example_data = [
            ['Eiche', 'hartholz', '1,3', 'Verfügbar', 'true'],
            ['Buche', 'hartholz', '1,2', 'Verfügbar', 'true'],
            ['Kiefer', 'nadelholz', '0,9', 'Standard', 'true'],
            ['Fichte', 'nadelholz', '0,8', 'Standard', 'true'],
            ['Nussbaum', 'hartholz', '1,8', 'Knapp', 'true'],
        ]

        if file_format == 'xlsx':
            return self._generate_excel(
                headers=headers,
                example_data=example_data,
                sheet_name='Holzarten',
                validations={
                    'kategorie': ['hartholz', 'weichholz', 'nadelholz'],
                    'verfuegbarkeit': ['Standard', 'Verfügbar', 'Knapp', 'Nicht verfügbar'],
                    'is_enabled': ['true', 'false']
                }
            )
        else:
            return self._generate_csv(headers, example_data)

    # =========================================================================
    # OBERFLÄCHENBEARBEITUNG (SURFACE FINISHES) TEMPLATE
    # =========================================================================

    def generate_oberflaechenbearbeitung_template(self, file_format: str = 'xlsx') -> bytes:
        """
        Generate template for OberflächenbearbeitungKennzahl bulk upload.

        Columns:
        - bearbeitung: Finish type name
        - preis_faktor: Price multiplier
        - zeit_faktor: Time multiplier
        - is_enabled: true/false
        """
        headers = ['bearbeitung', 'preis_faktor', 'zeit_faktor', 'is_enabled']

        example_data = [
            ['Naturbelassen', '1,0', '1,0', 'true'],
            ['Geölt', '1,10', '1,15', 'true'],
            ['Lackiert', '1,15', '1,25', 'true'],
            ['Gewachst', '1,05', '1,10', 'true'],
            ['Klavierlack', '1,6', '2,0', 'true'],
        ]

        if file_format == 'xlsx':
            return self._generate_excel(
                headers=headers,
                example_data=example_data,
                sheet_name='Oberflächen',
                validations={
                    'is_enabled': ['true', 'false']
                }
            )
        else:
            return self._generate_csv(headers, example_data)

    # =========================================================================
    # KOMPLEXITÄT (COMPLEXITY) TEMPLATE
    # =========================================================================

    def generate_komplexitaet_template(self, file_format: str = 'xlsx') -> bytes:
        """
        Generate template for KomplexitaetKennzahl bulk upload.

        Columns:
        - technik: Technique name
        - schwierigkeitsgrad: einfach/mittel/schwer/meister
        - preis_faktor: Price multiplier
        - zeit_faktor: Time multiplier
        - is_enabled: true/false
        """
        headers = ['technik', 'schwierigkeitsgrad', 'preis_faktor', 'zeit_faktor', 'is_enabled']

        example_data = [
            ['Sägen', 'einfach', '1,0', '1,0', 'true'],
            ['Gefräst', 'mittel', '1,15', '1,2', 'true'],
            ['Gedrechselt', 'schwer', '1,25', '1,5', 'true'],
            ['Geschnitzt', 'schwer', '1,5', '2,0', 'true'],
            ['Hand geschnitzt', 'meister', '2,0', '3,0', 'true'],
            ['Intarsien', 'meister', '2,5', '4,0', 'true'],
        ]

        if file_format == 'xlsx':
            return self._generate_excel(
                headers=headers,
                example_data=example_data,
                sheet_name='Komplexität',
                validations={
                    'schwierigkeitsgrad': ['einfach', 'mittel', 'schwer', 'meister'],
                    'is_enabled': ['true', 'false']
                }
            )
        else:
            return self._generate_csv(headers, example_data)

    # =========================================================================
    # MATERIALLISTE (MATERIAL CATALOG) TEMPLATE
    # =========================================================================

    def generate_materialliste_template(self, file_format: str = 'xlsx') -> bytes:
        """
        Generate template for MateriallistePosition bulk upload.

        Columns:
        - material_name: Material name
        - sku: Stock keeping unit (unique)
        - lieferant: Supplier name
        - standardkosten_eur: Standard cost (German format)
        - verpackungseinheit: Packaging unit
        - verfuegbarkeit: Auf Lager/Bestellbar/Nicht verfügbar
        - rabatt_ab_100: Discount % at 100+ qty (optional)
        - rabatt_ab_500: Discount % at 500+ qty (optional)
        - is_enabled: true/false
        """
        headers = [
            'material_name',
            'sku',
            'lieferant',
            'standardkosten_eur',
            'verpackungseinheit',
            'verfuegbarkeit',
            'rabatt_ab_100',
            'rabatt_ab_500',
            'is_enabled'
        ]

        example_data = [
            ['Schrauben 4x40mm', 'SCH-4x40-100', 'Würth GmbH', '12,50', 'Box zu 100 Stück', 'Auf Lager', '5', '10', 'true'],
            ['Leim PU D4', 'LEIM-D4-5L', 'Titebond', '45,80', '5 Liter Flasche', 'Auf Lager', '0', '15', 'true'],
            ['Schleifpapier K120', 'SCHLEIF-K120-50', 'Mirka', '28,90', 'Box zu 50 Blatt', 'Bestellbar', '10', '20', 'true'],
            ['Beschlag Topfband', 'BESCH-TOPF-35', 'Blum', '3,45', 'Stück', 'Auf Lager', '0', '0', 'true'],
        ]

        if file_format == 'xlsx':
            return self._generate_excel(
                headers=headers,
                example_data=example_data,
                sheet_name='Materialliste',
                validations={
                    'verfuegbarkeit': ['Auf Lager', 'Bestellbar', 'Nicht verfügbar'],
                    'is_enabled': ['true', 'false']
                }
            )
        else:
            return self._generate_csv(headers, example_data)

    # =========================================================================
    # SAISONALE MARGE (SEASONAL CAMPAIGNS) TEMPLATE
    # =========================================================================

    def generate_saisonale_marge_template(self, file_format: str = 'xlsx') -> bytes:
        """
        Generate template for SaisonaleMarge bulk upload.

        Columns:
        - name: Campaign name
        - description: Campaign description
        - adjustment_type: prozent/absolut
        - value: Adjustment value
        - start_date: DD.MM.YYYY
        - end_date: DD.MM.YYYY
        - applicable_to: Alle/Neukunden/Stammkunden/VIP
        - is_active: true/false
        """
        headers = [
            'name',
            'description',
            'adjustment_type',
            'value',
            'start_date',
            'end_date',
            'applicable_to',
            'is_active'
        ]

        # Generate example dates (current year)
        today = date.today()
        summer_start = date(today.year, 6, 1)
        summer_end = date(today.year, 8, 31)
        winter_start = date(today.year, 11, 1)
        winter_end = date(today.year + 1, 1, 31)

        example_data = [
            [
                'Sommeraktion 2025',
                'Rabatt für Neukunden im Sommer',
                'prozent',
                '10',
                summer_start.strftime('%d.%m.%Y'),
                summer_end.strftime('%d.%m.%Y'),
                'Neukunden',
                'true'
            ],
            [
                'Winterkampagne',
                'Winterrabatt für alle Kunden',
                'prozent',
                '5',
                winter_start.strftime('%d.%m.%Y'),
                winter_end.strftime('%d.%m.%Y'),
                'Alle',
                'true'
            ],
        ]

        if file_format == 'xlsx':
            return self._generate_excel(
                headers=headers,
                example_data=example_data,
                sheet_name='Saisonale Marge',
                validations={
                    'adjustment_type': ['prozent', 'absolut'],
                    'applicable_to': ['Alle', 'Neukunden', 'Stammkunden', 'VIP'],
                    'is_active': ['true', 'false']
                }
            )
        else:
            return self._generate_csv(headers, example_data)

    # =========================================================================
    # EXCEL GENERATION
    # =========================================================================

    def _generate_excel(
        self,
        headers: List[str],
        example_data: List[List[str]],
        sheet_name: str,
        validations: Dict[str, List[str]] = None
    ) -> bytes:
        """
        Generate Excel file with formatting and data validation.

        Args:
            headers: Column headers
            example_data: Example data rows
            sheet_name: Worksheet name
            validations: Dict of column_name -> [allowed_values] for dropdowns

        Returns:
            Excel file bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write example data
        for row_idx, row_data in enumerate(example_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = self.EXAMPLE_FILL

        # Add data validations (dropdowns)
        if validations:
            for col_name, allowed_values in validations.items():
                if col_name not in headers:
                    continue

                col_idx = headers.index(col_name) + 1
                col_letter = openpyxl.utils.get_column_letter(col_idx)

                # Create dropdown validation
                dv = DataValidation(
                    type='list',
                    formula1=f'"{",".join(allowed_values)}"',
                    allow_blank=False
                )
                dv.error = 'Ungültiger Wert'
                dv.errorTitle = 'Eingabefehler'
                dv.prompt = f'Wählen Sie: {", ".join(allowed_values)}'
                dv.promptTitle = 'Zulässige Werte'

                # Apply to column (rows 2-1000)
                ws.add_data_validation(dv)
                dv.add(f'{col_letter}2:{col_letter}1000')

        # Auto-size columns
        for col_idx, header in enumerate(headers, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            # Estimate width based on header length and data
            max_length = len(header)
            for row_data in example_data:
                if col_idx - 1 < len(row_data):
                    max_length = max(max_length, len(str(row_data[col_idx - 1])))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # =========================================================================
    # CSV GENERATION
    # =========================================================================

    def _generate_csv(
        self,
        headers: List[str],
        example_data: List[List[str]]
    ) -> bytes:
        """
        Generate CSV file.

        Args:
            headers: Column headers
            example_data: Example data rows

        Returns:
            CSV file bytes
        """
        output = io.StringIO()
        writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

        # Write headers
        writer.writerow(headers)

        # Write example data
        for row_data in example_data:
            writer.writerow(row_data)

        # Convert to bytes with UTF-8 BOM for Excel compatibility
        csv_bytes = output.getvalue().encode('utf-8-sig')
        return csv_bytes
